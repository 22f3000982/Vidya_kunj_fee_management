"""
Student Fee Management System - Backend
A Flask application to manage student fee records using Google Sheets as database.
"""

from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename
import pandas as pd
import os
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import gspread

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Configuration
DATA_FOLDER = 'data'
CREDENTIALS_FILE = 'credentials.json'  # Your Google service account JSON
SPREADSHEET_ID = '19F9qbeUSWyia-oWQbIWonJccytEUArW0ZrZ7kgmB0jc'  # Your Google Sheet ID

# Ensure data folder exists
os.makedirs(DATA_FOLDER, exist_ok=True)

# Google Sheets connection
gc = None
sheet = None

def get_google_sheet():
    """Connect to Google Sheet and return the worksheet"""
    global gc, sheet
    try:
        if gc is None:
            gc = gspread.service_account(filename=CREDENTIALS_FILE)
        if sheet is None:
            spreadsheet = gc.open_by_key(SPREADSHEET_ID)
            sheet = spreadsheet.sheet1
        return sheet
    except Exception as e:
        print(f"Error connecting to Google Sheets: {e}")
        return None


def read_sheet_data():
    """Read student data from Google Sheet"""
    try:
        worksheet = get_google_sheet()
        if worksheet is None:
            return []
        
        # Get all records as list of dicts
        records = worksheet.get_all_records()
        
        # Convert to proper format
        for record in records:
            for key in record:
                if record[key] is None:
                    record[key] = ''
                else:
                    record[key] = str(record[key])
        
        return records
    except Exception as e:
        print(f"Error reading Google Sheet: {e}")
        return []


def save_sheet_data(records):
    """Save all data to Google Sheet (clear and rewrite) - USE SPARINGLY!"""
    try:
        worksheet = get_google_sheet()
        if worksheet is None:
            return False
        
        # Clear the sheet
        worksheet.clear()
        
        if not records:
            # Write header only
            worksheet.append_row(['Student ID', 'Student Name', 'Father Name', 'Mobile Number', 'Month', 'Fee Status', 'Receipt Number'])
            return True
        
        # Write header
        headers = ['Student ID', 'Student Name', 'Father Name', 'Mobile Number', 'Month', 'Fee Status', 'Receipt Number']
        
        # Build all rows at once
        all_rows = [headers]
        for record in records:
            row = [
                str(record.get('Student ID', '')),
                str(record.get('Student Name', '')),
                str(record.get('Father Name', '')),
                str(record.get('Mobile Number', '')),
                str(record.get('Month', '')),
                str(record.get('Fee Status', '')),
                str(record.get('Receipt Number', ''))
            ]
            all_rows.append(row)
        
        # BATCH WRITE - Write all rows at once (MUCH FASTER!)
        worksheet.update('A1', all_rows)
        
        return True
    except Exception as e:
        print(f"Error saving to Google Sheet: {e}")
        return False


def update_row_in_sheet(row_number, record):
    """Update a specific row in Google Sheet (FAST - single API call)"""
    try:
        worksheet = get_google_sheet()
        if worksheet is None:
            return False
        
        row = [
            str(record.get('Student ID', '')),
            str(record.get('Student Name', '')),
            str(record.get('Father Name', '')),
            str(record.get('Mobile Number', '')),
            str(record.get('Month', '')),
            str(record.get('Fee Status', '')),
            str(record.get('Receipt Number', ''))
        ]
        
        # Update only the specific row (row_number is 1-indexed, +1 for header)
        worksheet.update(f'A{row_number}:G{row_number}', [row])
        return True
    except Exception as e:
        print(f"Error updating row in Google Sheet: {e}")
        return False


def delete_row_in_sheet(row_number):
    """Delete a specific row in Google Sheet (FAST - single API call)"""
    try:
        worksheet = get_google_sheet()
        if worksheet is None:
            return False
        
        worksheet.delete_rows(row_number)
        return True
    except Exception as e:
        print(f"Error deleting row in Google Sheet: {e}")
        return False


def append_rows_to_sheet(records):
    """Append multiple records to Google Sheet in one batch (FAST)"""
    try:
        worksheet = get_google_sheet()
        if worksheet is None:
            return False
        
        rows = []
        for record in records:
            row = [
                str(record.get('Student ID', '')),
                str(record.get('Student Name', '')),
                str(record.get('Father Name', '')),
                str(record.get('Mobile Number', '')),
                str(record.get('Month', '')),
                str(record.get('Fee Status', '')),
                str(record.get('Receipt Number', ''))
            ]
            rows.append(row)
        
        # BATCH APPEND - Much faster than individual appends
        worksheet.append_rows(rows)
        return True
    except Exception as e:
        print(f"Error appending rows to Google Sheet: {e}")
        return False


def find_row_number(student_name, father_name, month):
    """Find the row number for a specific record (1-indexed, includes header)"""
    try:
        worksheet = get_google_sheet()
        if worksheet is None:
            return None
        
        all_values = worksheet.get_all_values()
        
        for idx, row in enumerate(all_values):
            if idx == 0:  # Skip header
                continue
            # Row format: [Student ID, Student Name, Father Name, Mobile Number, Month, Fee Status, Receipt Number]
            if len(row) >= 5:
                if (str(row[1]).strip() == str(student_name).strip() and 
                    str(row[2]).strip() == str(father_name).strip() and
                    str(row[4]).strip().lower() == str(month).strip().lower()):
                    return idx + 1  # Return 1-indexed row number
        
        return None
    except Exception as e:
        print(f"Error finding row: {e}")
        return None


def append_to_sheet(record):
    """Append a single record to Google Sheet"""
    try:
        worksheet = get_google_sheet()
        if worksheet is None:
            return False
        
        row = [
            str(record.get('Student ID', '')),
            str(record.get('Student Name', '')),
            str(record.get('Father Name', '')),
            str(record.get('Mobile Number', '')),
            str(record.get('Month', '')),
            str(record.get('Fee Status', '')),
            str(record.get('Receipt Number', ''))
        ]
        worksheet.append_row(row)
        return True
    except Exception as e:
        print(f"Error appending to Google Sheet: {e}")
        return False


# Alias functions for compatibility with existing code
def read_excel_data():
    """Alias for read_sheet_data for compatibility"""
    return read_sheet_data()


def save_excel_data(records):
    """Alias for save_sheet_data for compatibility"""
    return save_sheet_data(records)


def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx', 'xls'}


@app.route('/')
def index():
    """Render main page"""
    return render_template('index.html')


@app.route('/api/students', methods=['GET'])
def get_students():
    """Get all student records with optional pagination"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    
    records = read_excel_data()
    total = len(records)
    
    # If pagination is requested
    if page and per_page:
        start = (page - 1) * per_page
        end = start + per_page
        paginated_records = records[start:end]
        total_pages = (total + per_page - 1) // per_page  # Ceiling division
        
        return jsonify({
            'success': True,
            'data': paginated_records,
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': total_pages,
            'has_next': page < total_pages,
            'has_prev': page > 1
        })
    
    # Return all records (for backward compatibility)
    return jsonify({
        'success': True,
        'data': records,
        'total': total
    })


@app.route('/api/search', methods=['GET'])
def search_students():
    """Search students by name, father name, or receipt number with pagination"""
    query = request.args.get('query', '').strip().lower()
    month = request.args.get('month', '').strip().lower()
    status = request.args.get('status', '').strip().lower()
    receipt = request.args.get('receipt', '').strip().lower()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    
    records = read_excel_data()
    filtered = []
    
    for record in records:
        # Apply filters
        match = True
        
        # If searching by receipt number specifically
        if receipt:
            receipt_match = receipt in str(record.get('Receipt Number', '')).lower()
            if not receipt_match:
                match = False
        
        if query and match:
            name_match = query in str(record.get('Student Name', '')).lower()
            father_match = query in str(record.get('Father Name', '')).lower()
            # Also check if query matches receipt number
            receipt_query_match = query in str(record.get('Receipt Number', '')).lower()
            if not (name_match or father_match or receipt_query_match):
                match = False
        
        if month and match:
            if month not in str(record.get('Month', '')).lower():
                match = False
        
        if status and match:
            record_status = str(record.get('Fee Status', '')).lower()
            if status == 'paid' and record_status != 'paid':
                match = False
            elif status == 'unpaid' and record_status not in ['not paid', 'unpaid', 'pending']:
                match = False
        
        if match:
            filtered.append(record)
    
    total = len(filtered)
    
    # Apply pagination
    if page and per_page:
        start = (page - 1) * per_page
        end = start + per_page
        paginated = filtered[start:end]
        total_pages = (total + per_page - 1) // per_page
        
        return jsonify({
            'success': True,
            'data': paginated,
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': total_pages,
            'has_next': page < total_pages,
            'has_prev': page > 1
        })
    
    return jsonify({
        'success': True,
        'data': filtered,
        'total': total
    })


@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Upload a new Excel file"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file provided'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'error': 'Invalid file type. Please upload an Excel file (.xlsx or .xls)'}), 400
    
    try:
        # Read the uploaded file first to validate it
        df = pd.read_excel(file)
        
        # Convert all column names to strings first (in case any are datetime objects)
        df.columns = [str(col) if not isinstance(col, str) else col for col in df.columns]
        
        # Check if this is a horizontal format (months as columns)
        # Horizontal format has columns like: Student Name, Father Name, Jan-26, Dec-25, etc.
        # Vertical format has columns like: Student Name, Father Name, Month, Fee Status, Receipt Number
        
        required_vertical_cols = ['Month', 'Fee Status', 'Receipt Number']
        has_vertical_format = any(col.lower().strip() in ['month', 'fee status', 'receipt number'] 
                                  for col in df.columns)
        
        if not has_vertical_format:
            # This is horizontal format - convert to vertical
            df = convert_horizontal_to_vertical(df)
        else:
            # Standard vertical format processing
            column_mapping = {}
            standard_columns = {
                'student id': 'Student ID',
                'id': 'Student ID',
                'roll no': 'Student ID',
                'student name': 'Student Name',
                'name': 'Student Name',
                'father name': 'Father Name',
                'father\'s name': 'Father Name',
                'father': 'Father Name',
                'mobile': 'Mobile Number',
                'mobile number': 'Mobile Number',
                'phone': 'Mobile Number',
                'phone number': 'Mobile Number',
                'contact': 'Mobile Number',
                'month': 'Month',
                'fee status': 'Fee Status',
                'status': 'Fee Status',
                'receipt number': 'Receipt Number',
                'receipt': 'Receipt Number',
                'receipt no': 'Receipt Number'
            }
            
            for actual_col in df.columns:
                col_lower = str(actual_col).lower().strip()
                if col_lower in standard_columns:
                    column_mapping[actual_col] = standard_columns[col_lower]
            
            df = df.rename(columns=column_mapping)
            
            required_columns = ['Student ID', 'Student Name', 'Father Name', 'Mobile Number', 'Month', 'Fee Status', 'Receipt Number']
            for col in required_columns:
                if col not in df.columns:
                    df[col] = ''
            
            df = df[required_columns]
        
        # Fill NaN values
        df = df.fillna('')
        
        # Convert all values to strings, handling datetime objects
        for col in df.columns:
            df[col] = df[col].apply(lambda x: x.strftime('%B %Y') if hasattr(x, 'strftime') else (str(x) if x != '' and pd.notna(x) else ''))
        
        # Save the cleaned data
        df.to_excel(EXCEL_FILE, index=False)
        
        # Get the record count
        records = df.to_dict('records')
        
        return jsonify({
            'success': True,
            'message': f'File uploaded successfully! {len(records)} records found.',
            'total': len(records)
        })
    except Exception as e:
        print(f"Upload error: {e}")
        return jsonify({'success': False, 'error': f'Failed to process Excel file: {str(e)}'}), 500


def convert_horizontal_to_vertical(df):
    """
    Convert horizontal Excel format (months as columns) to vertical format.
    Input format: Student Name, Father Name, Jan-26, Dec-25, Nov-25...
    Output format: Student Name, Father Name, Month, Fee Status, Receipt Number
    """
    import re
    
    # Identify student info columns and month columns
    student_cols = []
    month_cols = []
    
    # Month patterns to detect
    month_patterns = [
        r'^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[-/]?\d{2,4}$',
        r'^(january|february|march|april|may|june|july|august|september|october|november|december)\s*\d{2,4}$',
        r'^\d{1,2}[-/]\d{2,4}$'
    ]
    
    month_abbr_to_full = {
        'jan': 'January', 'feb': 'February', 'mar': 'March', 'apr': 'April',
        'may': 'May', 'jun': 'June', 'jul': 'July', 'aug': 'August',
        'sep': 'September', 'oct': 'October', 'nov': 'November', 'dec': 'December'
    }
    
    for col in df.columns:
        col_str = str(col).lower().strip()
        is_month = False
        for pattern in month_patterns:
            if re.match(pattern, col_str):
                is_month = True
                break
        
        if is_month:
            month_cols.append(col)
        else:
            student_cols.append(col)
    
    # Map student columns to standard names
    col_mapping = {}
    for col in student_cols:
        col_lower = str(col).lower().strip()
        if 'name' in col_lower and 'father' not in col_lower:
            col_mapping[col] = 'Student Name'
        elif 'father' in col_lower:
            col_mapping[col] = 'Father Name'
    
    # Build vertical records
    records = []
    for _, row in df.iterrows():
        student_name = ''
        father_name = ''
        
        for col in student_cols:
            if col in col_mapping:
                val = str(row[col]) if pd.notna(row[col]) else ''
                if col_mapping[col] == 'Student Name':
                    student_name = val
                elif col_mapping[col] == 'Father Name':
                    father_name = val
        
        # Process each month column
        for month_col in month_cols:
            cell_value = str(row[month_col]) if pd.notna(row[month_col]) else ''
            
            # Parse the month column name (e.g., "Jan-26" -> "January 2026")
            month_str = str(month_col).strip()
            month_match = re.match(r'^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[-/]?(\d{2,4})$', 
                                   month_str.lower())
            if month_match:
                month_abbr = month_match.group(1)
                year = month_match.group(2)
                if len(year) == 2:
                    year = '20' + year
                month_full = f"{month_abbr_to_full[month_abbr]} {year}"
            else:
                month_full = month_str
            
            # Parse cell value: "Paid (RCP-011-JAN26)" or "Not Paid"
            if cell_value.lower().startswith('paid'):
                fee_status = 'Paid'
                # Extract receipt number from parentheses
                receipt_match = re.search(r'\(([^)]+)\)', cell_value)
                receipt_number = receipt_match.group(1) if receipt_match else ''
            else:
                fee_status = 'Not Paid'
                receipt_number = ''
            
            records.append({
                'Student Name': student_name,
                'Father Name': father_name,
                'Month': month_full,
                'Fee Status': fee_status,
                'Receipt Number': receipt_number
            })
    
    return pd.DataFrame(records)


@app.route('/api/quick-mark-paid', methods=['POST'])
def quick_mark_paid():
    """Quick mark a student's fee as paid with auto-generated receipt"""
    data = request.json
    
    if not data:
        return jsonify({'success': False, 'error': 'No data provided'}), 400
    
    student_name = data.get('student_name')
    father_name = data.get('father_name')
    month = data.get('month')
    
    if not all([student_name, month]):
        return jsonify({'success': False, 'error': 'Missing required fields'}), 400
    
    records = read_excel_data()
    
    # Generate auto receipt number: RCP-MMYY-XXX
    from datetime import datetime
    now = datetime.now()
    prefix = f"RCP-{now.strftime('%m%y')}-"
    
    # Find the highest receipt number with this prefix
    max_num = 0
    for record in records:
        receipt = str(record.get('Receipt Number', ''))
        if receipt.startswith(prefix):
            try:
                num = int(receipt.replace(prefix, ''))
                max_num = max(max_num, num)
            except:
                pass
    
    receipt_number = f"{prefix}{str(max_num + 1).zfill(3)}"
    
    # Find the row number for smart update
    row_number = find_row_number(student_name, father_name, month)
    
    if row_number:
        # Find the existing record to preserve other fields
        updated_record = None
        for record in records:
            if (str(record.get('Student Name', '')) == str(student_name) and 
                str(record.get('Father Name', '')) == str(father_name) and
                str(record.get('Month', '')).lower() == str(month).lower()):
                updated_record = record.copy()
                updated_record['Fee Status'] = 'Paid'
                updated_record['Receipt Number'] = receipt_number
                break
        
        if updated_record and update_row_in_sheet(row_number, updated_record):
            return jsonify({
                'success': True, 
                'message': 'Marked as paid!',
                'receipt_number': receipt_number
            })
        else:
            return jsonify({'success': False, 'error': 'Failed to save changes'}), 500
    else:
        return jsonify({'success': False, 'error': 'Record not found'}), 404


@app.route('/api/defaulters', methods=['GET'])
def get_defaulters():
    """Get list of students with pending fees for 1+ months"""
    min_months = request.args.get('min_months', 1, type=int)
    
    records = read_excel_data()
    
    # Group unpaid records by student (name + father name)
    student_unpaid = {}
    student_info = {}
    
    for record in records:
        if str(record.get('Fee Status', '')).lower() not in ['paid']:
            key = f"{record.get('Student Name', '')}_{record.get('Father Name', '')}"
            if key not in student_unpaid:
                student_unpaid[key] = []
                student_info[key] = {
                    'student_name': record.get('Student Name', ''),
                    'father_name': record.get('Father Name', ''),
                    'student_id': record.get('Student ID', ''),
                    'mobile_number': record.get('Mobile Number', '')
                }
            student_unpaid[key].append(record.get('Month', ''))
    
    # Filter students with min_months or more unpaid
    defaulters = []
    for key, months in student_unpaid.items():
        if len(months) >= min_months:
            info = student_info[key]
            defaulters.append({
                'student_name': info['student_name'],
                'father_name': info['father_name'],
                'student_id': info['student_id'],
                'mobile_number': info['mobile_number'],
                'unpaid_count': len(months),
                'unpaid_months': sorted(months)
            })
    
    # Sort by unpaid count (highest first)
    defaulters.sort(key=lambda x: x['unpaid_count'], reverse=True)
    
    return jsonify({
        'success': True,
        'defaulters': defaulters,
        'total': len(defaulters),
        'min_months': min_months
    })


@app.route('/api/update', methods=['POST'])
def update_record():
    """Update a student's fee status - OPTIMIZED with smart row update"""
    data = request.json
    
    if not data:
        return jsonify({'success': False, 'error': 'No data provided'}), 400
    
    student_name = data.get('student_name')
    father_name = data.get('father_name')
    month = data.get('month')
    fee_status = data.get('fee_status')
    receipt_number = data.get('receipt_number', '')
    
    if not all([student_name, month, fee_status]):
        return jsonify({'success': False, 'error': 'Missing required fields'}), 400
    
    records = read_excel_data()
    
    # Check for duplicate receipt number if provided (excluding current record)
    if receipt_number and fee_status.lower() == 'paid':
        for record in records:
            if (str(record.get('Receipt Number', '')).lower() == str(receipt_number).lower() and
                not (str(record.get('Student Name', '')) == str(student_name) and 
                     str(record.get('Father Name', '')) == str(father_name) and
                     str(record.get('Month', '')).lower() == str(month).lower())):
                return jsonify({'success': False, 'error': 'This receipt number already exists for another record'}), 400
    
    # Find the row number for smart update
    row_number = find_row_number(student_name, father_name, month)
    
    if row_number:
        # Build updated record
        # First find the existing record to preserve other fields
        updated_record = None
        for record in records:
            if (str(record.get('Student Name', '')) == str(student_name) and 
                str(record.get('Father Name', '')) == str(father_name) and
                str(record.get('Month', '')).lower() == str(month).lower()):
                updated_record = record.copy()
                updated_record['Fee Status'] = fee_status
                updated_record['Receipt Number'] = receipt_number if fee_status.lower() == 'paid' else ''
                break
        
        if updated_record and update_row_in_sheet(row_number, updated_record):
            return jsonify({'success': True, 'message': 'Record updated successfully!'})
        else:
            return jsonify({'success': False, 'error': 'Failed to update record'}), 500
    else:
        return jsonify({'success': False, 'error': 'Record not found'}), 404


@app.route('/api/bulk-add', methods=['POST'])
def bulk_add_records():
    """Add multiple student fee records at once - OPTIMIZED with batch append"""
    data = request.json
    
    if not data or 'records' not in data:
        return jsonify({'success': False, 'error': 'No records provided'}), 400
    
    new_records = data['records']
    
    if not new_records or len(new_records) == 0:
        return jsonify({'success': False, 'error': 'Empty records list'}), 400
    
    existing_records = read_excel_data()
    
    # Build lookup sets for duplicate checking
    existing_student_months = set()
    existing_receipts = set()
    
    for record in existing_records:
        key = f"{record.get('Student Name', '')}_{record.get('Father Name', '')}_{record.get('Month', '').lower()}"
        existing_student_months.add(key)
        receipt = str(record.get('Receipt Number', '')).lower().strip()
        if receipt:
            existing_receipts.add(receipt)
    
    added_count = 0
    skipped_count = 0
    errors = []
    records_to_add = []  # Collect records for batch insert
    
    for i, rec in enumerate(new_records):
        # Validate required fields
        student_id = rec.get('Student ID', '').strip()
        student_name = rec.get('Student Name', '').strip()
        father_name = rec.get('Father Name', '').strip()
        mobile_number = rec.get('Mobile Number', '').strip()
        month = rec.get('Month', '').strip()
        fee_status = rec.get('Fee Status', 'Not Paid').strip()
        receipt_number = rec.get('Receipt Number', '').strip()
        
        if not all([student_name, month]):
            errors.append(f"Row {i+1}: Missing required fields")
            skipped_count += 1
            continue
        
        # Check for duplicate student+month
        key = f"{student_name}_{father_name}_{month.lower()}"
        if key in existing_student_months:
            errors.append(f"Row {i+1}: {student_name} already has record for {month}")
            skipped_count += 1
            continue
        
        # Check for duplicate receipt number
        if receipt_number and receipt_number.lower() in existing_receipts:
            errors.append(f"Row {i+1}: Receipt {receipt_number} already exists")
            skipped_count += 1
            continue
        
        # Add the record to batch list
        new_record = {
            'Student ID': student_id,
            'Student Name': student_name,
            'Father Name': father_name,
            'Mobile Number': mobile_number,
            'Month': month,
            'Fee Status': fee_status,
            'Receipt Number': receipt_number
        }
        
        records_to_add.append(new_record)
        existing_student_months.add(key)
        if receipt_number:
            existing_receipts.add(receipt_number.lower())
        added_count += 1
    
    if added_count > 0:
        # BATCH INSERT - Much faster than individual appends!
        if append_rows_to_sheet(records_to_add):
            message = f'Added {added_count} records successfully!'
            if skipped_count > 0:
                message += f' ({skipped_count} skipped due to duplicates)'
            return jsonify({
                'success': True, 
                'message': message,
                'added': added_count,
                'skipped': skipped_count,
                'errors': errors[:10]  # Limit error messages
            })
        else:
            return jsonify({'success': False, 'error': 'Failed to save records'}), 500
    else:
        return jsonify({
            'success': False, 
            'error': f'No records added. {skipped_count} records skipped.',
            'errors': errors[:10]
        }), 400


@app.route('/api/add', methods=['POST'])
def add_record():
    """Add a new student fee record"""
    data = request.json
    
    if not data:
        return jsonify({'success': False, 'error': 'No data provided'}), 400
    
    required_fields = ['student_name', 'father_name', 'month']
    for field in required_fields:
        if not data.get(field):
            return jsonify({'success': False, 'error': f'Missing required field: {field}'}), 400
    
    records = read_excel_data()
    
    # Check for duplicate entry (same student name + father name + month)
    for record in records:
        if (str(record.get('Student Name', '')) == str(data['student_name']) and 
            str(record.get('Father Name', '')) == str(data['father_name']) and
            str(record.get('Month', '')).lower() == str(data['month']).lower()):
            return jsonify({'success': False, 'error': 'Record already exists for this student and month'}), 400
    
    # Check for duplicate receipt number if provided
    if data.get('receipt_number'):
        for record in records:
            if str(record.get('Receipt Number', '')).lower() == str(data['receipt_number']).lower():
                return jsonify({'success': False, 'error': 'This receipt number already exists'}), 400
    
    new_record = {
        'Student ID': data.get('student_id', ''),
        'Student Name': data['student_name'],
        'Father Name': data['father_name'],
        'Mobile Number': data.get('mobile_number', ''),
        'Month': data['month'],
        'Fee Status': data.get('fee_status', 'Not Paid'),
        'Receipt Number': data.get('receipt_number', '')
    }
    
    records.append(new_record)
    
    if save_excel_data(records):
        return jsonify({'success': True, 'message': 'Record added successfully!'})
    else:
        return jsonify({'success': False, 'error': 'Failed to save record'}), 500


@app.route('/api/delete', methods=['POST'])
def delete_record():
    """Delete a student fee record - OPTIMIZED with direct row delete"""
    data = request.json
    
    student_name = data.get('student_name')
    father_name = data.get('father_name')
    month = data.get('month')
    
    if not all([student_name, month]):
        return jsonify({'success': False, 'error': 'Missing required fields'}), 400
    
    # Find the row number for direct delete
    row_number = find_row_number(student_name, father_name, month)
    
    if row_number:
        if delete_row_in_sheet(row_number):
            return jsonify({'success': True, 'message': 'Record deleted successfully!'})
        else:
            return jsonify({'success': False, 'error': 'Failed to delete record'}), 500
    else:
        return jsonify({'success': False, 'error': 'Record not found'}), 404


@app.route('/api/download', methods=['GET'])
def download_file():
    """Download the Excel file in a clean pivoted format"""
    filter_type = request.args.get('filter', 'all').lower()  # all, paid, unpaid
    records = read_excel_data()
    
    if not records:
        return jsonify({'success': False, 'error': 'No data to download'}), 404
    
    # Filter records based on filter_type
    if filter_type == 'paid':
        records = [r for r in records if str(r.get('Fee Status', '')).lower() == 'paid']
    elif filter_type == 'unpaid':
        records = [r for r in records if str(r.get('Fee Status', '')).lower() in ['not paid', 'unpaid', 'pending', '']]
    
    if not records:
        filter_label = 'paid' if filter_type == 'paid' else 'unpaid'
        return jsonify({'success': False, 'error': f'No {filter_label} records to download'}), 404
    
    try:
        # Get unique students and months
        students = {}
        all_months = set()
        
        for record in records:
            # Create unique key from name + father
            student_key = f"{record.get('Student Name', '')}_{record.get('Father Name', '')}"
            month = record.get('Month', '')
            
            if student_key not in students:
                students[student_key] = {
                    'Student Name': record.get('Student Name', ''),
                    'Father Name': record.get('Father Name', ''),
                    'months': {}
                }
            
            # Store fee status with receipt
            fee_status = record.get('Fee Status', '')
            receipt = record.get('Receipt Number', '')
            
            if fee_status.lower() == 'paid' and receipt:
                students[student_key]['months'][month] = f"Paid ({receipt})"
            elif fee_status.lower() == 'paid':
                students[student_key]['months'][month] = "Paid"
            else:
                students[student_key]['months'][month] = "Not Paid"
            
            all_months.add(month)
        
        # Sort months (most recent first)
        def month_sort_key(month_str):
            try:
                parts = month_str.split()
                month_names = ['January', 'February', 'March', 'April', 'May', 'June', 
                              'July', 'August', 'September', 'October', 'November', 'December']
                month_idx = month_names.index(parts[0]) if parts[0] in month_names else 0
                year = int(parts[1]) if len(parts) > 1 else 2026
                return (year, month_idx)
            except:
                return (0, 0)
        
        sorted_months = sorted(all_months, key=month_sort_key, reverse=True)
        
        # Build export data
        export_data = []
        for student_key, student_info in students.items():
            row = {
                'Student Name': student_info['Student Name'],
                'Father Name': student_info['Father Name']
            }
            for month in sorted_months:
                row[month] = student_info['months'].get(month, '-')
            export_data.append(row)
        
        # Sort by student name
        export_data.sort(key=lambda x: x['Student Name'])
        
        # Create DataFrame with ordered columns
        columns = ['Student Name', 'Father Name'] + sorted_months
        df = pd.DataFrame(export_data, columns=columns)
        
        # Save to temporary file
        export_file = os.path.join(DATA_FOLDER, 'export_temp.xlsx')
        
        # Use ExcelWriter for better formatting
        with pd.ExcelWriter(export_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Fee Records')
            
            worksheet = writer.sheets['Fee Records']
            
            # Define styles
            red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
            header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            red_font = Font(color='CC0000', bold=True)
            green_font = Font(color='006600')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Style header row
            for col_idx, col in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Style data cells and highlight Not Paid in red
            for row_idx in range(2, len(df) + 2):
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Check if it's a month column (after first 2 columns)
                    if col_idx > 2:
                        cell_value = str(cell.value) if cell.value else ''
                        if 'Not Paid' in cell_value:
                            cell.fill = red_fill
                            cell.font = red_font
                        elif 'Paid' in cell_value:
                            cell.fill = green_fill
                            cell.font = green_font
            
            # Auto-adjust column widths
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).map(len).max() if len(df) > 0 else 0,
                    len(str(col))
                ) + 2
                col_letter = chr(65 + idx) if idx < 26 else 'A' + chr(65 + idx - 26)
                worksheet.column_dimensions[col_letter].width = min(max_length, 30)
            
            # Set row height for header
            worksheet.row_dimensions[1].height = 25
        
        return send_file(
            export_file,
            as_attachment=True,
            download_name=f'student_fees_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    except Exception as e:
        print(f"Error creating export: {e}")
        return jsonify({'success': False, 'error': 'Failed to create export file'}), 500


@app.route('/api/summary', methods=['GET'])
def get_summary():
    """Get summary statistics"""
    records = read_excel_data()
    
    total = len(records)
    paid = sum(1 for r in records if str(r.get('Fee Status', '')).lower() == 'paid')
    unpaid = total - paid
    
    # Get unique months
    months = list(set(r.get('Month', '') for r in records if r.get('Month')))
    
    return jsonify({
        'success': True,
        'summary': {
            'total': total,
            'paid': paid,
            'unpaid': unpaid,
            'months': sorted(months)
        }
    })


@app.route('/api/unique-students', methods=['GET'])
def get_unique_students():
    """Get unique list of students (name + father name) for autocomplete"""
    records = read_excel_data()
    
    # Get unique student combinations
    seen = set()
    students = []
    for record in records:
        name = str(record.get('Student Name', '')).strip()
        father = str(record.get('Father Name', '')).strip()
        key = f"{name}_{father}"
        if key not in seen and name:
            seen.add(key)
            students.append({
                'name': name,
                'father': father,
                'student_id': str(record.get('Student ID', '')).strip(),
                'mobile': str(record.get('Mobile Number', '')).strip(),
                'display': f"{name} (F: {father})"
            })
    
    # Sort alphabetically by name
    students.sort(key=lambda x: x['name'].lower())
    
    return jsonify({
        'success': True,
        'students': students
    })


@app.route('/api/student/<receipt_number>', methods=['GET'])
def get_student_by_receipt(receipt_number):
    """Get student profile by receipt number"""
    records = read_excel_data()
    
    # Find the record with matching receipt number
    for record in records:
        if str(record.get('Receipt Number', '')).lower() == receipt_number.lower():
            student_name = record.get('Student Name', '')
            father_name = record.get('Father Name', '')
            # Get all records for this student
            student_records = [r for r in records if 
                str(r.get('Student Name', '')) == str(student_name) and
                str(r.get('Father Name', '')) == str(father_name)]
            
            # Calculate payment summary
            total_months = len(student_records)
            paid_months = sum(1 for r in student_records if str(r.get('Fee Status', '')).lower() == 'paid')
            
            return jsonify({
                'success': True,
                'student': {
                    'name': record.get('Student Name', ''),
                    'father_name': record.get('Father Name', ''),
                    'total_months': total_months,
                    'paid_months': paid_months,
                    'unpaid_months': total_months - paid_months
                },
                'records': student_records
            })
    
    return jsonify({'success': False, 'error': 'Receipt number not found'}), 404


@app.route('/api/update-student-profile', methods=['POST'])
def update_student_profile():
    """Update student profile info across all their records"""
    data = request.json
    
    if not data:
        return jsonify({'success': False, 'error': 'No data provided'}), 400
    
    original_name = data.get('original_name', '').strip()
    original_father = data.get('original_father', '').strip()
    new_student_id = data.get('student_id', '').strip()
    new_name = data.get('student_name', '').strip()
    new_father = data.get('father_name', '').strip()
    new_mobile = data.get('mobile_number', '').strip()
    
    if not new_name:
        return jsonify({'success': False, 'error': 'Student name is required'}), 400
    
    records = read_excel_data()
    updated_count = 0
    
    for record in records:
        if (str(record.get('Student Name', '')) == original_name and 
            str(record.get('Father Name', '')) == original_father):
            record['Student ID'] = new_student_id
            record['Student Name'] = new_name
            record['Father Name'] = new_father
            record['Mobile Number'] = new_mobile
            updated_count += 1
    
    if updated_count == 0:
        return jsonify({'success': False, 'error': 'No records found for this student'}), 404
    
    if save_excel_data(records):
        return jsonify({
            'success': True,
            'updated': updated_count,
            'message': f'Updated {updated_count} records'
        })
    else:
        return jsonify({'success': False, 'error': 'Failed to save changes'}), 500


@app.route('/api/student-profile/<path:student_key>', methods=['GET'])
def get_student_profile(student_key):
    """Get complete student profile with all payment records"""
    records = read_excel_data()
    
    # Parse student key (name_father)
    parts = student_key.split('_')
    student_name = parts[0] if len(parts) > 0 else ''
    father_name = parts[1] if len(parts) > 1 else ''
    
    # Get all records for this student
    student_records = [r for r in records if 
        str(r.get('Student Name', '')) == str(student_name) and
        str(r.get('Father Name', '')) == str(father_name)]
    
    if not student_records:
        return jsonify({'success': False, 'error': 'Student not found'}), 404
    
    # Get student info from first record
    first_record = student_records[0]
    
    # Calculate payment summary
    total_months = len(student_records)
    paid_months = sum(1 for r in student_records if str(r.get('Fee Status', '')).lower() == 'paid')
    
    return jsonify({
        'success': True,
        'student': {
            'name': first_record.get('Student Name', ''),
            'father_name': first_record.get('Father Name', ''),
            'total_months': total_months,
            'paid_months': paid_months,
            'unpaid_months': total_months - paid_months
        },
        'records': student_records
    })


if __name__ == '__main__':
    print("=" * 50)
    print("  Student Fee Management System")
    print("  Starting server at http://localhost:5000")
    print("=" * 50)
    app.run(debug=True, port=5000)
